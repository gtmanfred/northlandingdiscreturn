import io
from dataclasses import dataclass, field
from datetime import date, datetime
import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession
from app.phone import normalize_phone
from app.repositories.disc import DiscRepository
from app.repositories.owner import OwnerRepository
from app.services.welcome import maybe_enqueue_welcome
from app.services.heads_up import maybe_enqueue_heads_up

SHEET_NAME = "Current"
HEADER_KEYWORD = "Name"


@dataclass
class ParsedDiscRow:
    row_number: int
    first_name: str
    last_name: str
    phone: str | None
    manufacturer: str
    model: str
    colors: list[str]
    notes: str | None
    input_date: date | None
    returned: bool
    returned_date: date | None
    error: str | None = None


def _split_name(raw) -> tuple[str, str]:
    text = (str(raw).strip() if raw is not None else "")
    if text in ("", "?"):
        return "", ""
    parts = text.split()
    if len(parts) == 1:
        return "", parts[0]
    return parts[0], " ".join(parts[1:])


def _as_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def parse_current_sheet(file_bytes: bytes) -> list[ParsedDiscRow]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError("Current sheet not found")
    ws = wb[SHEET_NAME]
    grid = list(ws.iter_rows(values_only=True))

    header_idx = next(
        (i for i, r in enumerate(grid)
         if r and r[0] and HEADER_KEYWORD in str(r[0])),
        None,
    )
    if header_idx is None:
        raise ValueError("Header row not found in Current sheet")

    rows: list[ParsedDiscRow] = []
    for offset, r in enumerate(grid[header_idx + 1:], start=header_idx + 2):
        cells = list(r) + [None] * (10 - len(r))
        name, phone, mfr, model, color, other, code, found, returned_dt, _ = cells[:10]

        mfr = (str(mfr).strip() if mfr else "")
        model = (str(model).strip() if model else "")
        if not mfr and not model:
            continue  # blank row

        first, last = _split_name(name)
        phone_norm = None
        if phone:
            try:
                phone_norm = normalize_phone(str(phone))
            except ValueError:
                phone_norm = None

        colors = [c.lower() for c in str(color).split()] if color else []
        input_date = _as_date(found)
        ret_date = _as_date(returned_dt)
        code_str = (str(code).strip().upper() if code else "")
        returned = ret_date is not None or code_str == "R"
        if returned and ret_date is None:
            ret_date = input_date

        error = None
        if input_date is None:
            error = "missing or invalid Date found"

        rows.append(ParsedDiscRow(
            row_number=offset,
            first_name=first,
            last_name=last,
            phone=phone_norm,
            manufacturer=mfr,
            model=model,
            colors=colors,
            notes=(str(other).strip() if other else None),
            input_date=input_date,
            returned=returned,
            returned_date=ret_date if returned else None,
            error=error,
        ))
    return rows


def _compute_updates(existing, row: "ParsedDiscRow", owner_id) -> dict:
    """Fields that would change on an existing disc for this row. Empty dict = unchanged."""
    updates: dict = {}
    if (existing.notes or None) != (row.notes or None):
        updates["notes"] = row.notes
    if [c.strip().lower() for c in existing.colors] != [c.strip().lower() for c in row.colors]:
        updates["colors"] = row.colors
    if existing.owner_id != owner_id:
        updates["owner_id"] = owner_id
    # one-way return: only ever set returned, never clear
    if row.returned and not existing.is_returned:
        updates["is_returned"] = True
        updates["returned_date"] = row.returned_date
    return updates


@dataclass
class ImportSummary:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    errors: list[dict] = field(default_factory=list)


async def apply_import(rows: list[ParsedDiscRow], db: AsyncSession) -> ImportSummary:
    summary = ImportSummary()
    disc_repo = DiscRepository(db)
    owner_repo = OwnerRepository(db)

    for row in rows:
        if row.error or row.input_date is None:
            summary.errors.append({"row": row.row_number, "reason": row.error or "no date found"})
            continue

        owner_id = None
        owner_obj = None
        if row.phone or row.first_name or row.last_name:
            owner_obj = await owner_repo.resolve_or_create(
                first_name=row.first_name,
                last_name=row.last_name,
                phone_number=row.phone,
            )
            owner_id = owner_obj.id

        existing = await disc_repo.find_by_import_key(
            input_date=row.input_date,
            manufacturer=row.manufacturer,
            name=row.model,
            colors=row.colors,
            phone=row.phone,
        )

        if existing is None:
            disc = await disc_repo.create(
                manufacturer=row.manufacturer,
                name=row.model,
                colors=row.colors,
                input_date=row.input_date,
                owner_id=owner_id,
                notes=row.notes,
            )
            if row.returned:
                await disc_repo.update(
                    disc, is_returned=True, returned_date=row.returned_date
                )
            else:
                if owner_obj is not None:
                    await maybe_enqueue_welcome(owner=owner_obj, db=db)
                    await maybe_enqueue_heads_up(owner=owner_obj, disc=disc, db=db)
            summary.created += 1
        else:
            updates = _compute_updates(existing, row, owner_id)
            if updates:
                await disc_repo.update(existing, **updates)
                summary.updated += 1
            else:
                summary.skipped += 1

    await db.flush()
    return summary


def row_to_dict(r: ParsedDiscRow) -> dict:
    return {
        "row_number": r.row_number,
        "first_name": r.first_name,
        "last_name": r.last_name,
        "phone": r.phone,
        "manufacturer": r.manufacturer,
        "model": r.model,
        "colors": r.colors,
        "notes": r.notes,
        "input_date": r.input_date.isoformat() if r.input_date else None,
        "returned": r.returned,
        "returned_date": r.returned_date.isoformat() if r.returned_date else None,
        "error": r.error,
    }


def row_from_dict(d: dict) -> ParsedDiscRow:
    return ParsedDiscRow(
        row_number=d["row_number"],
        first_name=d["first_name"],
        last_name=d["last_name"],
        phone=d["phone"],
        manufacturer=d["manufacturer"],
        model=d["model"],
        colors=d["colors"],
        notes=d["notes"],
        input_date=date.fromisoformat(d["input_date"]) if d["input_date"] else None,
        returned=d["returned"],
        returned_date=date.fromisoformat(d["returned_date"]) if d["returned_date"] else None,
        error=d["error"],
    )


def _owner_label_from_row(row: ParsedDiscRow) -> str | None:
    name = f"{row.first_name} {row.last_name}".strip()
    parts = [p for p in (name, row.phone) if p]
    return " / ".join(parts) if parts else None


def _owner_label(owner) -> str | None:
    if owner is None:
        return None
    name = f"{owner.first_name} {owner.last_name}".strip()
    parts = [p for p in (name, owner.phone_number) if p]
    return " / ".join(parts) if parts else None


def _disc_label(row: ParsedDiscRow) -> dict:
    return {
        "manufacturer": row.manufacturer,
        "model": row.model,
        "colors": row.colors,
        "owner": _owner_label_from_row(row),
    }


def _plan_diffs(existing, row: ParsedDiscRow) -> list[dict]:
    """Human-readable field changes for display. Semantic (does not use owner_id)."""
    diffs: list[dict] = []
    if (existing.notes or None) != (row.notes or None):
        diffs.append({"field": "notes", "old": existing.notes, "new": row.notes})
    if [c.strip().lower() for c in existing.colors] != [c.strip().lower() for c in row.colors]:
        diffs.append({"field": "colors", "old": existing.colors, "new": row.colors})
    row_has_owner = bool(row.phone or row.first_name or row.last_name)
    old_owner = _owner_label(existing.owner)
    new_owner = _owner_label_from_row(row) if row_has_owner else None
    if old_owner != new_owner:
        diffs.append({"field": "owner", "old": old_owner, "new": new_owner})
    if row.returned and not existing.is_returned:
        diffs.append({"field": "returned", "old": False, "new": True})
    return diffs


@dataclass
class ImportPlan:
    created: list[dict] = field(default_factory=list)
    updated: list[dict] = field(default_factory=list)
    unchanged: int = 0
    errors: list[dict] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "created": self.created,
            "updated": self.updated,
            "unchanged": self.unchanged,
            "errors": self.errors,
            "counts": {
                "created": len(self.created),
                "updated": len(self.updated),
                "unchanged": self.unchanged,
                "errors": len(self.errors),
            },
        }


async def plan_import(rows: list[ParsedDiscRow], db: AsyncSession) -> ImportPlan:
    """Read-only classification of what an import would do. No writes, no SMS."""
    disc_repo = DiscRepository(db)
    plan = ImportPlan()
    for row in rows:
        if row.error or row.input_date is None:
            plan.errors.append(
                {"row": row_to_dict(row), "reason": row.error or "no date found"}
            )
            continue
        existing = await disc_repo.find_by_import_key(
            input_date=row.input_date,
            manufacturer=row.manufacturer,
            name=row.model,
            colors=row.colors,
            phone=row.phone,
        )
        label = {"row_number": row.row_number, **_disc_label(row)}
        if existing is None:
            plan.created.append(label)
        else:
            diffs = _plan_diffs(existing, row)
            if diffs:
                plan.updated.append({**label, "diffs": diffs})
            else:
                plan.unchanged += 1
    return plan
