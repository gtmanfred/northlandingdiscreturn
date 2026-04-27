"""Import discs and owners from the North Landing Found Disc spreadsheet via the HTTP API.

Usage (from repo root):
    export NLD_API_KEY=hou_...   # admin API key (POST /users/me/api-key)
    uv run --project backend python scripts/import_discs.py <xlsx_path> [options]

Options:
    --base-url URL      Backend base URL. Default: $NLD_API_URL or http://localhost:8000.
    --api-key KEY       Bearer API key. Default: $NLD_API_KEY.
    --sheet NAME        Sheet to import (repeatable). Default: Current, No Number, Returned.
    --dry-run           Parse and report; do not call the API.
    --limit N           Stop after N rows per sheet (useful for smoke tests).

Spreadsheet columns (data starts at row 4):
    Name | Phone | Mfr | Model | Color | Other | Code | Date found | Date returned | Date contacted

Behavior:
    * Every non-empty row is imported. Missing string fields become "" and a missing
      Date found falls back to today (the API requires it).
    * Owner is sent only when Name is a real name (not blank / '?') AND Phone normalizes
      to a US 10-digit number. Otherwise the disc is created ownerless.
    * is_returned cannot be set on POST /discs; if a Date returned is present we follow up
      with PATCH /discs/{id} {"is_returned": true}.
    * is_clear is inferred from the color text ("clear", "trans", or "tint").
    * No client-side dedup: the API has no exact-match lookup. Re-running creates duplicates.
      Use --dry-run first to sanity-check counts.
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import httpx
from openpyxl import load_workbook

DEFAULT_SHEETS = ["Current", "No Number", "Returned"]
HEADER_ROW = 3  # 0-indexed; data begins at row index 3 (row 4 in Excel).
DEFAULT_BASE_URL = "http://localhost:8000"


@dataclass
class ParsedRow:
    name: str | None
    phone: str | None
    manufacturer: str
    model: str
    color: str
    date_found: date
    date_found_missing: bool
    date_returned: date | None
    is_clear: bool


def _clean(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _to_date(value) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def _try_phone(raw: str | None) -> str | None:
    """Return E.164 +1XXXXXXXXXX if `raw` resolves to a 10-digit US number, else None."""
    if not raw:
        return None
    digits = re.sub(r"\D", "", raw)
    if len(digits) == 10:
        return f"+1{digits}"
    if len(digits) == 11 and digits[0] == "1":
        return f"+{digits}"
    return None


def _is_real_name(raw: str | None) -> bool:
    if not raw:
        return False
    return bool(raw.strip().strip("?").strip())


def parse_row(row: tuple) -> ParsedRow | None:
    name = _clean(row[0])
    phone_raw = _clean(row[1])
    manufacturer = _clean(row[2])
    model = _clean(row[3])
    color = _clean(row[4])
    date_found = _to_date(row[7])
    date_returned = _to_date(row[8])

    if not any((name, phone_raw, manufacturer, model, color, date_found, date_returned)):
        return None

    color_lower = (color or "").lower()
    is_clear = any(token in color_lower for token in ("clear", "trans", "tint"))

    return ParsedRow(
        name=name if _is_real_name(name) else None,
        phone=_try_phone(phone_raw),
        manufacturer=manufacturer or "",
        model=model or "",
        color=color or "",
        date_found=date_found or date.today(),
        date_found_missing=date_found is None,
        date_returned=date_returned,
        is_clear=is_clear,
    )


def build_create_payload(parsed: ParsedRow) -> dict:
    payload = {
        "manufacturer": parsed.manufacturer,
        "name": parsed.model,
        "color": parsed.color,
        "input_date": parsed.date_found.isoformat(),
        "is_clear": parsed.is_clear,
        "is_found": True,
    }
    if parsed.name and parsed.phone:
        payload["owner_name"] = parsed.name
        payload["phone_number"] = parsed.phone
    return payload


def import_sheet(
    ws,
    *,
    client: httpx.Client | None,
    dry_run: bool,
    limit: int | None,
) -> dict:
    stats = {
        "rows": 0,
        "skipped": 0,
        "imported": 0,
        "owners": 0,
        "ownerless": 0,
        "missing_date": 0,
        "marked_returned": 0,
        "errors": 0,
    }

    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx <= HEADER_ROW:
            continue
        if limit is not None and stats["imported"] >= limit:
            break
        stats["rows"] += 1

        parsed = parse_row(row)
        if parsed is None:
            stats["skipped"] += 1
            continue
        if parsed.date_found_missing:
            stats["missing_date"] += 1

        payload = build_create_payload(parsed)
        has_owner = "owner_name" in payload
        if has_owner:
            stats["owners"] += 1
        else:
            stats["ownerless"] += 1

        if dry_run:
            stats["imported"] += 1
            if parsed.date_returned is not None:
                stats["marked_returned"] += 1
            continue

        try:
            resp = client.post("/discs", json=payload)
            resp.raise_for_status()
        except httpx.HTTPError as e:
            stats["errors"] += 1
            detail = ""
            if isinstance(e, httpx.HTTPStatusError):
                detail = f" body={e.response.text[:200]}"
            print(f"[error] row {idx + 1}: {e}{detail}", file=sys.stderr)
            continue

        stats["imported"] += 1
        disc_id = resp.json()["id"]

        if parsed.date_returned is not None:
            try:
                patch = client.patch(f"/discs/{disc_id}", json={"is_returned": True})
                patch.raise_for_status()
                stats["marked_returned"] += 1
            except httpx.HTTPError as e:
                stats["errors"] += 1
                detail = ""
                if isinstance(e, httpx.HTTPStatusError):
                    detail = f" body={e.response.text[:200]}"
                print(
                    f"[error] row {idx + 1} mark-returned (id={disc_id}): {e}{detail}",
                    file=sys.stderr,
                )

    return stats


def run(
    *,
    xlsx_path: Path,
    sheets: list[str],
    base_url: str,
    api_key: str | None,
    dry_run: bool,
    limit: int | None,
) -> int:
    wb = load_workbook(xlsx_path, data_only=True)

    if dry_run:
        client_cm = _NullClient()
    else:
        if not api_key:
            print(
                "error: API key required (set $NLD_API_KEY or pass --api-key)",
                file=sys.stderr,
            )
            return 2
        client_cm = httpx.Client(
            base_url=base_url,
            headers={"Authorization": f"Bearer {api_key}"},
            timeout=30.0,
        )

    totals = {
        "rows": 0,
        "skipped": 0,
        "imported": 0,
        "owners": 0,
        "ownerless": 0,
        "missing_date": 0,
        "marked_returned": 0,
        "errors": 0,
    }

    with client_cm as client:
        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                print(f"[warn] sheet not found: {sheet_name!r}", file=sys.stderr)
                continue
            print(f"=== {sheet_name} ===")
            stats = import_sheet(
                wb[sheet_name],
                client=client if not dry_run else None,
                dry_run=dry_run,
                limit=limit,
            )
            for k, v in stats.items():
                totals[k] += v
            print(f"  {stats}")

    print(f"\nTOTAL: {totals}")
    if dry_run:
        print("(dry-run: no API calls made)")
    return 1 if totals["errors"] else 0


class _NullClient:
    """Stand-in context manager for dry-run mode."""

    def __enter__(self):
        return None

    def __exit__(self, *exc):
        return False


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    p.add_argument("xlsx_path", type=Path)
    p.add_argument("--base-url", default=os.environ.get("NLD_API_URL", DEFAULT_BASE_URL))
    p.add_argument("--api-key", default=os.environ.get("NLD_API_KEY"))
    p.add_argument("--sheet", action="append", dest="sheets")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--limit", type=int)
    args = p.parse_args()

    if not args.xlsx_path.exists():
        print(f"file not found: {args.xlsx_path}", file=sys.stderr)
        return 2

    return run(
        xlsx_path=args.xlsx_path,
        sheets=args.sheets or DEFAULT_SHEETS,
        base_url=args.base_url,
        api_key=args.api_key,
        dry_run=args.dry_run,
        limit=args.limit,
    )


if __name__ == "__main__":
    raise SystemExit(main())
